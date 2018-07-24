import kivy
import openpyxl
import pprint
import os

# kivy imports for different widgets, layouts, etc...
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.properties import DictProperty
from kivy.properties import ObjectProperty
from kivy.properties import StringProperty
from kivy.properties import BooleanProperty
from kivy.config import Config
from kivy.uix.textinput import TextInput
from kivy.lang import Builder
from kivy.uix.filechooser import *
from kivy.uix.popup import Popup
from kivy.core.window import Window
from kivy.uix.filechooser import FileChooser

kivy.require("1.10.0")

Window.size = (500, 100)

class LoadDialog(FloatLayout):
    load = ObjectProperty(None)
    cancel = ObjectProperty(None)


# root element
class QuoTam(BoxLayout):
    file_path_input = ObjectProperty()
    search_results = DictProperty({})
    loadfile = ObjectProperty(None)
    text_input = StringProperty('')


    def dismiss_popup(self):
        self._popup.dismiss()

    def show_load(self):
        Window.size = (500, 400)
        content = LoadDialog(load=self.load, cancel=self.cancel)        #updates the variables in LoadDialog with the functions in this class.
        self._popup = Popup(title="Load file", content=content,
                            size_hint=(0.9, 0.9))
        self._popup.open()

    def load(self, path, filename):
        Window.size = (500, 100)
        self.file_path_input = os.path.join(path, filename[0])
        self.text_input = self.file_path_input

        self.dismiss_popup()

    def cancel(self):
        Window.size = (500, 100)

        self.dismiss_popup()

    # This function finds the coordinates (row: name) of all of the devices in the BOM
    def find_all_devices(self):
        # This is a list of all of the devices
        all_devices = {}
        dev_rows = {}

        # Import statement
        try:
            wb = openpyxl.load_workbook(self.file_path_input.strip('\"'))
            ws = wb.active
        except:
            wb = openpyxl.load_workbook(self.file_path_input.text.strip('\"'))
            ws = wb.active

        ''' Creates a dictionary with {deviceName: None} key-value pairs
        that will contain the names of all of the devices in the BOM.
        '''
        for i in range(1, ws.max_row):
            colC_Val = ws.cell(row=i, column=3).value
            all_devices[colC_Val] = None


            for device in all_devices.keys():
                if device == colC_Val and device != None:

                    dev_rows[i] = device

        ''' Above for loop creates a dictionary of {rows: deviceName} entries 
        that will be used to find device quantities in the next function. 
        '''

        things = [dev_rows, all_devices]
        return things


    def device_counter(self):

        # Import statement
        try:
            wb = openpyxl.load_workbook(self.file_path_input.strip('\"'))
            ws = wb.active
        except:
            wb = openpyxl.load_workbook(self.file_path_input.text.strip('\"'))
            ws = wb.active


        a = QuoTam.find_all_devices(self)
        row_nameDirectory = a[0]
        deviceNames = a[1]
        results = {}

        ''' These two for loops take the name of each individual device, 
        find all of the rows it exists in and come up with a total 
        for that specific device. 
        '''

        for dev in deviceNames.keys():
            device_total = 0

            for row, name in row_nameDirectory.items():
                if dev == name and type(ws.cell(row=row, column=2).value)==int:
                    device_total += ws.cell(row=row, column=2).value

            results[dev] = device_total



        ''' This statement creates/opens the .txt results file that will contain the device quantities. 
        Only devices that are NOT in the "dontCareLog.txt" file are included in this "resultsFile.txt" 
        '''

        results_file = open('resultsFile.txt', 'w')
        dc_file = open('dontCareLog.txt')
        dc_list = dc_file.read().split("\n,")

        results_file.write("Device Name: Quantity\n\n\n\n")

        for k,v in results.items():
            for dontcare in range(len(dc_list)):
                if str(k) not in dc_list[dontcare] and str(v) != '0':

                    results_file.write(str(k) + ": " + str(v) + "\n")

        os.startfile('resultsFile.txt')
        results_file.close()
        dc_file.close()

    # Terminate the program
    def cancel_button(self):
        exit()

    pass


# This class runs the LoginApp
class QuoTamApp(App):

  def build(self):
        return QuoTam()


if __name__ == '__main__':
    QuoTamApp().run()
